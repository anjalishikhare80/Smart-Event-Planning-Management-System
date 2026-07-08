from dotenv import load_dotenv
load_dotenv()


from flask import Flask, render_template, request, redirect, session, jsonify, send_file
from werkzeug.security import check_password_hash
import re
import os
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from io import BytesIO
from datetime import datetime
import hmac
import hashlib
import qrcode
import secrets
from flask_wtf.csrf import CSRFProtect
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.pdfgen import canvas
from reportlab.lib.units import inch




from openpyxl import Workbook

from database import (
    init_db,
    create_user, get_user_by_email, get_user_by_id,
    get_user_by_email_and_mobile, update_user_password,
    set_organizer_trusted, is_organizer_trusted,
    expire_events,
    get_events_needing_reminder, get_event_registrant_emails, mark_reminder_sent,
    create_event, get_event_by_id, check_duplicate_event,
    get_pending_events_by_organizer, get_rejected_events_by_organizer,
    get_my_events_by_organizer,
    update_pending_event, resubmit_event, delete_owned_event,
    close_registration, get_event_participants, get_organizer_stats,
    mark_attendance, submit_feedback, get_feedback_by_registration, get_event_feedback,
    get_visible_events_for_participants, get_all_events, admin_delete_event,
    get_pending_events_all, approve_event, reject_event,
    bulk_approve_events, bulk_reject_events,
    register_for_event, cancel_registration, get_user_registrations,
    get_registration_by_id, get_all_registrations_by_organizer,
    get_all_registrations_admin,
    get_all_users, get_admin_stats, delete_user,
    create_notification, get_notifications, get_unread_notification_count,
    mark_notification_read, mark_all_notifications_read,
    save_contact_message, get_contact_messages, mark_contact_read,
    get_contact_message_by_id, reply_to_contact_message,
)

from chatbot import chatbot_bp

app = Flask(__name__)

_env_secret = os.environ.get('SECRET_KEY')
if _env_secret:
    app.secret_key = _env_secret
else:
    # No hardcoded fallback: a known static secret would let anyone forge
    # session cookies. Generate a random one instead — sessions just won't
    # survive a restart until SECRET_KEY is set in .env.
    app.secret_key = secrets.token_hex(32)
    print("⚠️  SECRET_KEY not set in .env — using a random key for this run. "
          "Set SECRET_KEY in .env for stable sessions across restarts.")

app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
# Set SESSION_COOKIE_SECURE=True once this app is served over HTTPS in
# production (it must stay False for plain-HTTP local development, or
# cookies won't be sent at all).
app.config['SESSION_COOKIE_SECURE'] = os.environ.get('FLASK_ENV') == 'production'

csrf = CSRFProtect(app)

limiter = Limiter(
    app=app,
    key_func=get_remote_address,
    default_limits=[]
)
app.register_blueprint(chatbot_bp)
# The chatbot endpoint only returns a canned reply based on role/message —
# it doesn't read or write anything sensitive, so CSRF protection (which
# guards against unwanted state changes) doesn't apply here. Exempting it
# also avoids needing to plumb a CSRF token into the static chatbot.js file.
csrf.exempt(chatbot_bp)

# ── Email (Gmail SMTP) ────────────────────────────────────────────────────────
MAIL_SERVER   = 'smtp.gmail.com'
MAIL_PORT     = 587
MAIL_USERNAME = os.environ.get('MAIL_USERNAME', '')
MAIL_PASSWORD = os.environ.get('MAIL_APP_PASSWORD', '')
ADMIN_EMAIL   = 'eventhub62@gmail.com'

# ── Updated 6 event categories ────────────────────────────────────────────────
VALID_CATEGORIES = [
    'Technical Events', 'Creative Events', 'Cultural Events',
    'Academic Events', 'Competition', 'Sports & Fitness'
]

# ── Pre-defined rejection reasons ─────────────────────────────────────────────
REJECTION_REASONS = [
    'Incomplete event information',
    'Invalid or past event date',
    'Duplicate event already exists',
    'Venue details are unclear',
    'Event description is insufficient',
    'Capacity is unrealistic',
    'Event violates platform policies',
    'Contact information is missing',
]


# ============= HELPER FUNCTIONS ============= #

def validate_email(email):
    return re.match(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$',
                    email) is not None


def validate_mobile(mobile):
    return re.match(r'^\d{10}$', str(mobile)) is not None


def validate_password(password):
    return len(password) >= 6


def validate_event_dates(event_date, registration_close_date):
    try:
        event_dt = datetime.strptime(event_date, "%Y-%m-%d").date()
    except (ValueError, TypeError):
        return "Invalid event date"

    today = datetime.now().date()

    if event_dt < today:
        return "Event date cannot be in the past"

    if registration_close_date:
        try:
            close_dt = datetime.strptime(registration_close_date, "%Y-%m-%d").date()
        except ValueError:
            return "Invalid registration closing date"

        if close_dt < today:
            return "Registration closing date cannot be in the past"

        if close_dt > event_dt:
            return "Registration closing date must be on or before the event date"

    return None


def validate_event_fields(title, category, event_date, venue,
                           capacity, registration_close_date):
    if not all([title, category, event_date, venue, capacity]):
        return "All required fields must be filled in"

    if len(title.strip()) < 3:
        return "Event title must be at least 3 characters"

    if category not in VALID_CATEGORIES:
        return "Invalid event category selected"

    try:
        cap = int(capacity)
    except (ValueError, TypeError):
        return "Capacity must be a valid number"

    if cap <= 0:
        return "Capacity must be greater than 0"

    if cap > 100000:
        return "Capacity cannot exceed 100,000"

    return validate_event_dates(event_date, registration_close_date)


def is_user_logged_in():
    return 'user_id' in session


def get_user_dashboard_route(role):
    return {
        "Admin":       "/admin_dashboard",
        "Organizer":   "/organizer_dashboard",
        "Participant": "/participant_dashboard"
    }.get(role, "/login")


# ============= EMAIL HELPERS ============= #

def send_email(to_email, subject, html_body):
    if not MAIL_USERNAME or not MAIL_PASSWORD:
        print(f"⚠️  Email skipped (credentials not set) → {to_email} | {subject}")
        return False
    try:
        msg = MIMEMultipart('alternative')
        msg['Subject'] = subject
        msg['From']    = f"EventHub <{MAIL_USERNAME}>"
        msg['To']      = to_email
        msg.attach(MIMEText(html_body, 'html'))

        with smtplib.SMTP(MAIL_SERVER, MAIL_PORT) as srv:
            srv.ehlo()
            srv.starttls()
            srv.login(MAIL_USERNAME, MAIL_PASSWORD)
            srv.sendmail(MAIL_USERNAME, to_email, msg.as_string())

        print(f"✅ Email sent → {to_email} | {subject}")
        return True
    except Exception as e:
        print(f"❌ Email failed → {to_email} | {e}")
        return False


def _email_base(content_html):
    return f"""<!DOCTYPE html>
<html><head><meta charset="UTF-8"></head>
<body style="margin:0;padding:0;background:#0d0e15;font-family:Inter,sans-serif;">
<div style="max-width:600px;margin:32px auto;background:#181a24;border-radius:14px;
            overflow:hidden;border:1px solid rgba(255,255,255,0.1);">
  <div style="background:#f2b134;padding:16px 28px;">
    <span style="color:#18181c;font-weight:800;font-size:1.15rem;">🎫 EventHub</span>
  </div>
  <div style="padding:28px 28px 20px;">{content_html}</div>
  <div style="padding:16px 28px;border-top:1px solid rgba(255,255,255,0.08);
              font-size:0.75rem;color:#9a97aa;text-align:center;">
    This is an automated message from EventHub. Please do not reply.
  </div>
</div></body></html>"""


def notify_admin_new_event(event_title, organizer_name,
                            event_date, venue, event_code):
    content = f"""
      <h2 style="color:#F3F1ED;margin:0 0 8px;">New Event Pending Your Approval</h2>
      <p style="color:#9a97aa;margin:0 0 22px;font-size:0.9rem;">
        {organizer_name} submitted a new event for review.
      </p>
      <div style="background:#1f222e;border-radius:10px;padding:18px;margin-bottom:22px;">
        <table style="width:100%;border-collapse:collapse;">
          <tr><td style="padding:7px 0;color:#9a97aa;font-size:0.82rem;width:130px;">Event</td>
              <td style="padding:7px 0;color:#F3F1ED;font-weight:600;">{event_title}</td></tr>
          <tr><td style="padding:7px 0;color:#9a97aa;font-size:0.82rem;">Code</td>
              <td style="padding:7px 0;color:#f2b134;font-family:monospace;">{event_code}</td></tr>
          <tr><td style="padding:7px 0;color:#9a97aa;font-size:0.82rem;">Organizer</td>
              <td style="padding:7px 0;color:#F3F1ED;">{organizer_name}</td></tr>
          <tr><td style="padding:7px 0;color:#9a97aa;font-size:0.82rem;">Date</td>
              <td style="padding:7px 0;color:#F3F1ED;">{event_date}</td></tr>
          <tr><td style="padding:7px 0;color:#9a97aa;font-size:0.82rem;">Venue</td>
              <td style="padding:7px 0;color:#F3F1ED;">{venue}</td></tr>
        </table>
      </div>
      <p style="color:#9a97aa;font-size:0.85rem;margin:0;">
        Log in to your Admin Dashboard to approve or reject this event.
      </p>"""
    send_email(ADMIN_EMAIL,
               f"[EventHub] Pending Approval — {event_title}",
               _email_base(content))


def notify_organizer_approved(organizer_email, organizer_name,
                               event_title, event_code, event_date):
    content = f"""
      <h2 style="color:#57c2b6;margin:0 0 8px;">🎉 Your Event Has Been Approved!</h2>
      <p style="color:#9a97aa;margin:0 0 22px;font-size:0.9rem;">
        Great news, {organizer_name}! Your event is now live.
      </p>
      <div style="background:#1f222e;border-radius:10px;padding:18px;margin-bottom:22px;">
        <table style="width:100%;border-collapse:collapse;">
          <tr><td style="padding:7px 0;color:#9a97aa;font-size:0.82rem;width:130px;">Event</td>
              <td style="padding:7px 0;color:#F3F1ED;font-weight:600;">{event_title}</td></tr>
          <tr><td style="padding:7px 0;color:#9a97aa;font-size:0.82rem;">Code</td>
              <td style="padding:7px 0;color:#f2b134;font-family:monospace;">{event_code}</td></tr>
          <tr><td style="padding:7px 0;color:#9a97aa;font-size:0.82rem;">Date</td>
              <td style="padding:7px 0;color:#F3F1ED;">{event_date}</td></tr>
        </table>
      </div>
      <p style="color:#9a97aa;font-size:0.85rem;margin:0;">
        Participants can now discover and register for your event.
      </p>"""
    send_email(organizer_email,
               f"[EventHub] Event Approved — {event_title}",
               _email_base(content))


def notify_organizer_rejected(organizer_email, organizer_name,
                               event_title, reason):
    content = f"""
      <h2 style="color:#ef6361;margin:0 0 8px;">Event Requires Attention</h2>
      <p style="color:#9a97aa;margin:0 0 22px;font-size:0.9rem;">
        Hi {organizer_name}, your event needs changes before approval.
      </p>
      <div style="background:#1f222e;border-radius:10px;padding:18px;margin-bottom:22px;">
        <div style="color:#9a97aa;font-size:0.78rem;text-transform:uppercase;
                    letter-spacing:0.1em;margin-bottom:8px;">Event</div>
        <div style="color:#F3F1ED;font-weight:600;margin-bottom:16px;">{event_title}</div>
        <div style="color:#9a97aa;font-size:0.78rem;text-transform:uppercase;
                    letter-spacing:0.1em;margin-bottom:8px;">Rejection Reason</div>
        <div style="color:#ef6361;">{reason}</div>
      </div>
      <p style="color:#9a97aa;font-size:0.85rem;margin:0;">
        Please fix the issue and resubmit from your Organizer Dashboard.
      </p>"""
    send_email(organizer_email,
               f"[EventHub] Event Needs Changes — {event_title}",
               _email_base(content))


def notify_participant_registered(participant_email, participant_name,
                                   event_title, event_code, event_date,
                                   venue, registration_id):
    content = f"""
      <h2 style="color:#57c2b6;margin:0 0 8px;">✅ Registration Confirmed!</h2>
      <p style="color:#9a97aa;margin:0 0 22px;font-size:0.9rem;">
        Hi {participant_name}, you are registered for the following event.
      </p>
      <div style="background:#1f222e;border-radius:10px;padding:18px;margin-bottom:22px;">
        <table style="width:100%;border-collapse:collapse;">
          <tr><td style="padding:7px 0;color:#9a97aa;font-size:0.82rem;width:160px;">Registration ID</td>
              <td style="padding:7px 0;color:#f2b134;font-family:monospace;font-weight:700;">
                REG-{registration_id:05d}</td></tr>
          <tr><td style="padding:7px 0;color:#9a97aa;font-size:0.82rem;">Event</td>
              <td style="padding:7px 0;color:#F3F1ED;font-weight:600;">{event_title}</td></tr>
          <tr><td style="padding:7px 0;color:#9a97aa;font-size:0.82rem;">Event Code</td>
              <td style="padding:7px 0;color:#f2b134;font-family:monospace;">{event_code}</td></tr>
          <tr><td style="padding:7px 0;color:#9a97aa;font-size:0.82rem;">Date</td>
              <td style="padding:7px 0;color:#F3F1ED;">{event_date}</td></tr>
          <tr><td style="padding:7px 0;color:#9a97aa;font-size:0.82rem;">Venue</td>
              <td style="padding:7px 0;color:#F3F1ED;">{venue}</td></tr>
        </table>
      </div>
      <p style="color:#9a97aa;font-size:0.85rem;margin:0;">
        Download your registration receipt from your Participant Dashboard.
      </p>"""
    send_email(participant_email,
               f"[EventHub] Registration Confirmed — {event_title}",
               _email_base(content))


def notify_contact_confirmation(to_email, full_name, subject):
    """Send confirmation email to user after they submit a contact form."""
    content = f"""
      <h2 style="color:#57c2b6;margin:0 0 8px;">We Have Received Your Query</h2>
      <p style="color:#9a97aa;margin:0 0 22px;font-size:0.9rem;">
        Hi {full_name}, thank you for reaching out to EventHub.
      </p>
      <div style="background:#1f222e;border-radius:10px;padding:18px;margin-bottom:22px;">
        <div style="color:#9a97aa;font-size:0.78rem;text-transform:uppercase;
                    letter-spacing:0.1em;margin-bottom:8px;">Subject</div>
        <div style="color:#F3F1ED;font-weight:600;">{subject or 'General Enquiry'}</div>
      </div>
      <p style="color:#9a97aa;font-size:0.85rem;margin:0;">
        Our team will get back to you as soon as possible.
      </p>"""
    send_email(to_email,
               "[EventHub] We Have Received Your Query",
               _email_base(content))


def notify_contact_reply(to_email, full_name, subject, original_message, reply_text):
    """Send the admin's reply to the person who submitted the contact form."""
    content = f"""
      <h2 style="color:#f2b134;margin:0 0 8px;">Reply to Your Query</h2>
      <p style="color:#9a97aa;margin:0 0 22px;font-size:0.9rem;">
        Hi {full_name}, here's a response to your message.
      </p>
      <div style="background:#1f222e;border-radius:10px;padding:18px;margin-bottom:16px;">
        <div style="color:#9a97aa;font-size:0.78rem;text-transform:uppercase;
                    letter-spacing:0.1em;margin-bottom:8px;">Your message — {subject or 'General Enquiry'}</div>
        <div style="color:#9a97aa;font-size:0.85rem;font-style:italic;">{original_message}</div>
      </div>
      <div style="background:#22261f;border-left:3px solid #f2b134;border-radius:6px;padding:16px 18px;">
        <div style="color:#f2b134;font-size:0.78rem;text-transform:uppercase;
                    letter-spacing:0.1em;margin-bottom:8px;">EventHub Team's Reply</div>
        <div style="color:#F3F1ED;font-size:0.9rem;line-height:1.6;">{reply_text}</div>
      </div>"""
    send_email(to_email,
               "[EventHub] Reply to Your Query",
               _email_base(content))


def notify_event_reminder(to_email, full_name, event_title, event_date,
                          event_time, venue, event_code):
    content = f"""
      <h2 style="color:#f2b134;margin:0 0 8px;">⏰ Your Event Is Tomorrow!</h2>
      <p style="color:#9a97aa;margin:0 0 22px;font-size:0.9rem;">
        Hi {full_name}, just a friendly reminder about your upcoming event.
      </p>
      <div style="background:#1f222e;border-radius:10px;padding:18px;margin-bottom:22px;">
        <table style="width:100%;border-collapse:collapse;">
          <tr><td style="padding:7px 0;color:#9a97aa;font-size:0.82rem;width:130px;">Event</td>
              <td style="padding:7px 0;color:#F3F1ED;font-weight:600;">{event_title}</td></tr>
          <tr><td style="padding:7px 0;color:#9a97aa;font-size:0.82rem;">Date</td>
              <td style="padding:7px 0;color:#F3F1ED;">{event_date}{' · ' + event_time if event_time else ''}</td></tr>
          <tr><td style="padding:7px 0;color:#9a97aa;font-size:0.82rem;">Venue</td>
              <td style="padding:7px 0;color:#F3F1ED;">{venue}</td></tr>
          <tr><td style="padding:7px 0;color:#9a97aa;font-size:0.82rem;">Event Code</td>
              <td style="padding:7px 0;color:#f2b134;font-family:monospace;">{event_code}</td></tr>
        </table>
      </div>
      <p style="color:#9a97aa;font-size:0.85rem;margin:0;">
        Don't forget to bring your registration receipt or check-in QR code.
      </p>"""
    send_email(to_email,
               f"[EventHub] Reminder — {event_title} is tomorrow!",
               _email_base(content))


def send_event_reminders():
    """
    Emails everyone registered for events happening tomorrow, once per event
    (tracked via events.reminder_sent). Called opportunistically wherever
    expire_events() already runs — same lazy pattern this app already uses
    for date-based state changes, no background scheduler required.

    Caveat: since nothing runs this on a timer, a reminder only goes out
    once someone (any participant/organizer/admin) loads a dashboard page
    on the day before the event. For guaranteed delivery regardless of
    site traffic, this would need a real scheduled job (e.g. APScheduler
    or an OS-level cron hitting a dedicated endpoint) instead.
    """
    events = get_events_needing_reminder()
    for event in events:
        registrants = get_event_registrant_emails(event['id'])
        for person in registrants:
            notify_event_reminder(
                person['email'], person['fullname'], event['title'],
                event['event_date'], event['event_time'], event['venue'],
                event['event_code']
            )
        mark_reminder_sent(event['id'])


def _get_organizer_by_event(event):
    """Helper: return the organizer user row for a given event dict."""
    if not event:
        return None
    return get_user_by_id(event['organizer_id'])


# ============= REPORT HELPERS ============= #

def build_pdf_table(title, headers, rows):
    buffer = BytesIO()
    doc    = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    elements = [Paragraph(title, styles['Title']), Spacer(1, 16)]
    data = [headers] + [
        [str(c) if c is not None else '-' for c in row] for row in rows
    ]
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND',   (0, 0), (-1, 0), colors.HexColor('#11121a')),
        ('TEXTCOLOR',    (0, 0), (-1, 0), colors.white),
        ('FONTNAME',     (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',     (0, 0), (-1, -1), 8),
        ('GRID',         (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS',(0,1), (-1, -1), [colors.white, colors.HexColor('#f2f2f2')]),
        ('VALIGN',       (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING',   (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING',(0, 0), (-1, -1), 6),
        ('LEFTPADDING',  (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return buffer


def build_excel_table(headers, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    ws.append(headers)
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)
    for row in rows:
        ws.append([c if c is not None else '-' for c in row])
    for col_idx, header in enumerate(headers, start=1):
        col_values = [str(header)] + [str(r[col_idx - 1]) for r in rows]
        max_len    = max(len(v) for v in col_values)
        ws.column_dimensions[
            ws.cell(row=1, column=col_idx).column_letter
        ].width = max_len + 4
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def build_certificate_pdf(reg):
    """
    Builds a landscape 'Certificate of Participation' PDF for a completed
    event registration.
    """
    buffer = BytesIO()
    W, H = landscape(letter)
    c = canvas.Canvas(buffer, pagesize=(W, H))

    navy = colors.HexColor('#11121a')
    gold = colors.HexColor('#c9a34e')

    # Background + border
    c.setFillColor(colors.white)
    c.rect(0, 0, W, H, fill=1, stroke=0)

    c.setStrokeColor(gold)
    c.setLineWidth(3)
    c.rect(0.4 * inch, 0.4 * inch, W - 0.8 * inch, H - 0.8 * inch, fill=0, stroke=1)
    c.setStrokeColor(navy)
    c.setLineWidth(1)
    c.rect(0.5 * inch, 0.5 * inch, W - 1.0 * inch, H - 1.0 * inch, fill=0, stroke=1)

    # Header
    c.setFillColor(navy)
    c.setFont('Helvetica-Bold', 14)
    c.drawCentredString(W / 2, H - 1.1 * inch, "EVENTHUB")
    c.setFont('Helvetica', 10)
    c.setFillColor(colors.grey)
    c.drawCentredString(W / 2, H - 1.32 * inch, "Smart Event Planning & Management System")

    c.setFillColor(gold)
    c.setFont('Times-BoldItalic', 34)
    c.drawCentredString(W / 2, H - 2.15 * inch, "Certificate of Participation")

    c.setFillColor(navy)
    c.setFont('Helvetica', 13)
    c.drawCentredString(W / 2, H - 2.75 * inch, "This is to certify that")

    c.setFont('Times-BoldItalic', 26)
    c.drawCentredString(W / 2, H - 3.35 * inch, reg['participant_name'])

    c.setFont('Helvetica', 13)
    c.drawCentredString(W / 2, H - 3.85 * inch, "has successfully participated in")

    c.setFont('Helvetica-Bold', 18)
    c.drawCentredString(W / 2, H - 4.35 * inch, reg['event_title'])

    c.setFont('Helvetica', 12)
    detail_line = f"organized by {reg['organizer_name']}  ·  {reg['event_date']}"
    if reg.get('venue'):
        detail_line += f"  ·  {reg['venue']}"
    c.drawCentredString(W / 2, H - 4.75 * inch, detail_line)

    # Footer: codes + signature lines
    c.setFont('Helvetica', 9)
    c.setFillColor(colors.grey)
    footer_left = f"Reg. ID: REG-{reg['registration_id']:05d}"
    if reg.get('event_code'):
        footer_left += f"   ·   Event Code: {reg['event_code']}"
    c.drawString(0.9 * inch, 1.1 * inch, footer_left)
    c.drawString(0.9 * inch, 0.92 * inch,
                  f"Issued on {datetime.now().strftime('%d %b %Y')}")

    c.setStrokeColor(navy)
    c.line(W - 3.2 * inch, 1.15 * inch, W - 0.9 * inch, 1.15 * inch)
    c.setFillColor(navy)
    c.setFont('Helvetica', 9)
    c.drawCentredString(W - 2.05 * inch, 0.98 * inch, "Authorized Signatory — EventHub")

    c.showPage()
    c.save()
    buffer.seek(0)
    return buffer


def build_checkin_token(registration_id):
    """
    Short HMAC token proving a check-in link genuinely came from a QR code we
    generated (not just someone guessing/incrementing a registration id in
    the URL bar). Uses the app's SECRET_KEY, so it can't be forged without it.
    """
    return hmac.new(
        app.secret_key.encode(),
        str(registration_id).encode(),
        hashlib.sha256
    ).hexdigest()[:16]


def build_qr_png(data):
    """Builds a QR code PNG for the given string, returned as a BytesIO."""
    img = qrcode.make(data)
    buffer = BytesIO()
    img.save(buffer, format='PNG')
    buffer.seek(0)
    return buffer


# ============= HOME ============= #

@app.route('/')
def home():
    return render_template('home.html')


# ============= CONTACT FORM ============= #

@app.route('/contact', methods=['POST'])
def contact():
    """Store contact form submission and send emails."""
    full_name = request.form.get('full_name', '').strip()
    email     = request.form.get('email', '').strip().lower()
    user_role = request.form.get('user_role', '').strip()
    subject   = request.form.get('subject', '').strip()
    message   = request.form.get('message', '').strip()

    if not all([full_name, email, message]):
        return jsonify({'success': False,
                        'message': 'Name, email, and message are required'}), 400

    if not validate_email(email):
        return jsonify({'success': False,
                        'message': 'Invalid email format'}), 400

    try:
        save_contact_message(full_name, email, user_role, subject, message)

        # Email admin
        admin_content = f"""
          <h2 style="color:#F3F1ED;margin:0 0 8px;">New Contact Message</h2>
          <div style="background:#1f222e;border-radius:10px;padding:18px;margin-bottom:22px;">
            <table style="width:100%;border-collapse:collapse;">
              <tr><td style="padding:7px 0;color:#9a97aa;width:120px;">Name</td>
                  <td style="padding:7px 0;color:#F3F1ED;font-weight:600;">{full_name}</td></tr>
              <tr><td style="padding:7px 0;color:#9a97aa;">Email</td>
                  <td style="padding:7px 0;color:#F3F1ED;">{email}</td></tr>
              <tr><td style="padding:7px 0;color:#9a97aa;">Role</td>
                  <td style="padding:7px 0;color:#F3F1ED;">{user_role or '-'}</td></tr>
              <tr><td style="padding:7px 0;color:#9a97aa;">Subject</td>
                  <td style="padding:7px 0;color:#F3F1ED;">{subject or '-'}</td></tr>
              <tr><td style="padding:7px 0;color:#9a97aa;vertical-align:top;">Message</td>
                  <td style="padding:7px 0;color:#F3F1ED;">{message}</td></tr>
            </table>
          </div>"""
        send_email(ADMIN_EMAIL,
                   f"[EventHub] Contact Form — {full_name}",
                   _email_base(admin_content))

        # Confirmation to user
        notify_contact_confirmation(email, full_name, subject)

        return jsonify({'success': True,
                        'message': 'Your message has been sent! We will get back to you soon.'}), 200

    except Exception as e:
        print(f"❌ CONTACT ERROR: {e}")
        return jsonify({'success': False,
                        'message': 'Something went wrong. Please try again.'}), 500


# ============= REGISTER ============= #

@app.route('/register', methods=['GET', 'POST'])
@limiter.limit("10 per hour", methods=['POST'])
def register():
    if request.method == 'POST':
        fullname = request.form.get('fullname', '').strip()
        email    = request.form.get('email', '').strip().lower()
        mobile   = request.form.get('mobile', '').strip()
        role     = request.form.get('role', '').strip()
        password = request.form.get('password', '').strip()

        if not all([fullname, email, mobile, role, password]):
            return jsonify({'success': False,
                            'message': 'Please fill in all fields'}), 400

        if not validate_email(email):
            return jsonify({'success': False,
                            'message': 'Invalid email format'}), 400

        if not validate_mobile(mobile):
            return jsonify({'success': False,
                            'message': 'Mobile number must be exactly 10 digits'}), 400

        if not validate_password(password):
            return jsonify({'success': False,
                            'message': 'Password must be at least 6 characters'}), 400

        if role not in ["Organizer", "Participant"]:
            return jsonify({'success': False,
                            'message': 'Invalid role selected'}), 400

        try:
            created = create_user(fullname, email, mobile, role, password)
            if not created:
                return jsonify({'success': False,
                                'message': 'Email already registered.'}), 409

            # Notify admin of new organizer registration
            if role == 'Organizer':
                admin_user = get_user_by_email(ADMIN_EMAIL)
                if admin_user:
                    create_notification(
                        admin_user['id'],
                        "New Organizer Registered",
                        f"{fullname} ({email}) joined as an Organizer.",
                        'info'
                    )

            return jsonify({'success': True,
                            'message': 'Registration successful! Redirecting to login...',
                            'redirect': '/login'}), 201
        except Exception as e:
            print(f"❌ REGISTER ERROR: {e}")
            return jsonify({'success': False,
                            'message': 'An unexpected error occurred.'}), 500

    return redirect('/?auth=signup')


# ============= LOGIN ============= #

@app.route('/login', methods=['GET', 'POST'])
@limiter.limit("10 per minute", methods=['POST'])
def login():
    if request.method == 'POST':
        email    = request.form.get('email', '').strip().lower()
        password = request.form.get('password', '').strip()

        if not email or not password:
            return jsonify({'success': False,
                            'message': 'Please enter both email and password'}), 400

        if not validate_email(email):
            return jsonify({'success': False,
                            'message': 'Invalid email format'}), 400

        try:
            user = get_user_by_email(email)
            if user and check_password_hash(user['password'], password):
                session['user_id']   = user['id']
                session['user_name'] = user['fullname']
                session['email']     = user['email']
                session['role']      = user['role']

                return jsonify({'success': True,
                                'message': f'Welcome, {user["fullname"]}!',
                                'redirect': get_user_dashboard_route(user['role']),
                                'role': user['role']}), 200
            else:
                return jsonify({'success': False,
                                'message': 'Invalid email or password'}), 401

        except Exception as e:
            print(f"❌ LOGIN ERROR: {e}")
            return jsonify({'success': False,
                            'message': 'Login failed. Please try again.'}), 500

    return redirect('/?auth=login')


# ============= FORGOT / RESET PASSWORD ============= #

@app.route('/forgot_password', methods=['POST'])
def forgot_password():
    email  = request.form.get('email', '').strip().lower()
    mobile = request.form.get('mobile', '').strip()

    if not email or not mobile:
        return jsonify({'success': False,
                        'message': 'Please enter both email and mobile number'}), 400

    if not validate_email(email):
        return jsonify({'success': False, 'message': 'Invalid email format'}), 400

    if not validate_mobile(mobile):
        return jsonify({'success': False,
                        'message': 'Mobile number must be exactly 10 digits'}), 400

    try:
        user = get_user_by_email_and_mobile(email, mobile)
        if user:
            return jsonify({'success': True,
                            'message': 'Identity verified. Please set a new password.'}), 200
        return jsonify({'success': False,
                        'message': 'Invalid Email Address or Mobile Number.'}), 401
    except Exception as e:
        print(f"❌ FORGOT PASSWORD ERROR: {e}")
        return jsonify({'success': False,
                        'message': 'Something went wrong. Please try again.'}), 500


@app.route('/reset_password', methods=['POST'])
def reset_password():
    email            = request.form.get('email', '').strip().lower()
    mobile           = request.form.get('mobile', '').strip()
    new_password     = request.form.get('new_password', '').strip()
    confirm_password = request.form.get('confirm_password', '').strip()

    if not all([email, mobile, new_password, confirm_password]):
        return jsonify({'success': False, 'message': 'Please fill in all fields'}), 400

    if new_password != confirm_password:
        return jsonify({'success': False, 'message': 'Passwords do not match'}), 400

    if not validate_password(new_password):
        return jsonify({'success': False,
                        'message': 'Password must be at least 6 characters'}), 400

    try:
        user = get_user_by_email_and_mobile(email, mobile)
        if not user:
            return jsonify({'success': False,
                            'message': 'Invalid Email Address or Mobile Number.'}), 401

        update_user_password(email, new_password)
        return jsonify({'success': True,
                        'message': 'Password updated successfully. Please log in.',
                        'redirect': '/login'}), 200
    except Exception as e:
        print(f"❌ RESET PASSWORD ERROR: {e}")
        return jsonify({'success': False,
                        'message': 'Something went wrong. Please try again.'}), 500


# ============= NOTIFICATION API ============= #

@app.route('/notifications')
def get_notifications_route():
    """Return notifications JSON for the bell dropdown."""
    if not is_user_logged_in():
        return jsonify({'success': False}), 403

    user_id       = session['user_id']
    notifications = get_notifications(user_id, limit=15)
    unread_count  = get_unread_notification_count(user_id)

    return jsonify({
        'success':      True,
        'unread_count': unread_count,
        'notifications': [
            {
                'id':         n['id'],
                'title':      n['title'],
                'message':    n['message'],
                'type':       n['type'],
                'is_read':    n['is_read'],
                'created_at': str(n['created_at']),
            }
            for n in notifications
        ]
    }), 200


@app.route('/notifications/mark_read/<int:notif_id>', methods=['POST'])
def mark_notification_read_route(notif_id):
    if not is_user_logged_in():
        return jsonify({'success': False}), 403
    mark_notification_read(notif_id, session['user_id'])
    return jsonify({'success': True}), 200


@app.route('/notifications/mark_all_read', methods=['POST'])
def mark_all_read_route():
    if not is_user_logged_in():
        return jsonify({'success': False}), 403
    mark_all_notifications_read(session['user_id'])
    return jsonify({'success': True}), 200


# ============= DASHBOARDS ============= #

@app.route('/participant_dashboard')
def participant_dashboard():
    if not is_user_logged_in():
        return redirect('/login')
    if session.get('role') != 'Participant':
        return redirect('/')

    expire_events()
    send_event_reminders()

    visible_events   = get_visible_events_for_participants()
    my_registrations = get_user_registrations(session['user_id'])
    registered_ids   = {row['id'] for row in my_registrations}

    from datetime import date
    today = date.today().isoformat()

    discover_events = [
        e for e in visible_events
        if e['status'] == 'Approved'
        and e['registered_count'] < e['capacity']
        and (not e['registration_close_date'] or e['registration_close_date'] >= today)
    ]

    unread_count = get_unread_notification_count(session['user_id'])

    return render_template(
        'participant_dashboard.html',
        user_name=session.get('user_name'),
        email=session.get('email'),
        events=discover_events,
        my_registrations=my_registrations,
        registered_ids=registered_ids,
        unread_count=unread_count,
    )


@app.route('/organizer_dashboard')
def organizer_dashboard():
    if not is_user_logged_in():
        return redirect('/login')
    if session.get('role') != 'Organizer':
        return redirect('/')

    expire_events()
    send_event_reminders()
    organizer_id = session['user_id']
    unread_count = get_unread_notification_count(organizer_id)

    return render_template(
        'organizer_dashboard.html',
        user_name=session.get('user_name'),
        email=session.get('email'),
        stats=get_organizer_stats(organizer_id),
        events=get_my_events_by_organizer(organizer_id),
        pending_events=get_pending_events_by_organizer(organizer_id),
        rejected_events=get_rejected_events_by_organizer(organizer_id),
        all_registrations=get_all_registrations_by_organizer(organizer_id),
        is_trusted=is_organizer_trusted(organizer_id),
        unread_count=unread_count,
    )


@app.route('/admin_dashboard')
def admin_dashboard():
    if not is_user_logged_in():
        return redirect('/login')
    if session.get('role') != 'Admin':
        return redirect('/')

    expire_events()
    send_event_reminders()
    unread_count = get_unread_notification_count(session['user_id'])

    return render_template(
        'admin_dashboard.html',
        user_name=session.get('user_name'),
        email=session.get('email'),
        stats=get_admin_stats(),
        users=get_all_users(),
        events=get_all_events(),
        pending_events=get_pending_events_all(),
        all_registrations=get_all_registrations_admin(),
        rejection_reasons=REJECTION_REASONS,
        contact_messages=get_contact_messages(),
        unread_count=unread_count,
    )


# ============= ORGANIZER ACTIONS ============= #

def _extract_event_form(form):
    return {
        'title':                   form.get('title', '').strip(),
        'category':                form.get('category', '').strip(),
        'event_date':              form.get('event_date', '').strip(),
        'event_time':              form.get('event_time', '').strip(),
        'venue':                   form.get('venue', '').strip(),
        'description':             form.get('description', '').strip(),
        'capacity':                form.get('capacity', '').strip(),
        'registration_close_date': form.get('registration_close_date', '').strip(),
        'contact_number':          form.get('contact_number', '').strip(),
    }


@app.route('/organizer/create_event', methods=['POST'])
def create_event_route():
    if not is_user_logged_in() or session.get('role') != 'Organizer':
        return redirect('/login')

    f = _extract_event_form(request.form)

    error = validate_event_fields(
        f['title'], f['category'], f['event_date'],
        f['venue'], f['capacity'], f['registration_close_date']
    )
    if error:
        print(f"⚠️  Event validation failed: {error}")
        return redirect('/organizer_dashboard')

    if check_duplicate_event(f['title'], f['event_date'], f['venue']):
        print(f"⚠️  Duplicate event: {f['title']} on {f['event_date']} at {f['venue']}")
        return redirect('/organizer_dashboard')

    try:
        cap = int(f['capacity'])
    except ValueError:
        return redirect('/organizer_dashboard')

    event_id = create_event(
        f['title'], f['category'], f['event_date'], f['event_time'],
        f['venue'], f['description'], cap,
        f['registration_close_date'] or None,
        f['contact_number'], session['user_id']
    )

    new_event = get_event_by_id(event_id) if event_id else None

    if new_event:
        trusted = is_organizer_trusted(session['user_id'])

        if trusted:
            # Auto-approved — notify organizer
            create_notification(
                session['user_id'],
                "Event Auto-Approved",
                f"Your event '{new_event['title']}' was auto-approved and is now live.",
                'success'
            )
        else:
            # Needs admin review — notify admin
            admin_user = get_user_by_email(ADMIN_EMAIL)
            if admin_user:
                create_notification(
                    admin_user['id'],
                    "New Event Pending Approval",
                    f"{session.get('user_name')} submitted '{new_event['title']}' for review.",
                    'info'
                )
            notify_admin_new_event(
                new_event['title'], session.get('user_name', ''),
                new_event['event_date'], new_event['venue'],
                new_event['event_code'] or ''
            )

    return redirect('/organizer_dashboard')


@app.route('/organizer/edit_pending/<int:event_id>', methods=['POST'])
def edit_pending_event_route(event_id):
    if not is_user_logged_in() or session.get('role') != 'Organizer':
        return redirect('/login')

    f = _extract_event_form(request.form)
    error = validate_event_fields(
        f['title'], f['category'], f['event_date'],
        f['venue'], f['capacity'], f['registration_close_date']
    )
    if error:
        return redirect('/organizer_dashboard')

    try:
        cap = int(f['capacity'])
    except ValueError:
        return redirect('/organizer_dashboard')

    update_pending_event(
        event_id, session['user_id'], f['title'], f['category'],
        f['event_date'], f['event_time'], f['venue'], f['description'],
        cap, f['registration_close_date'] or None, f['contact_number']
    )
    return redirect('/organizer_dashboard')


@app.route('/organizer/resubmit/<int:event_id>', methods=['POST'])
def resubmit_event_route(event_id):
    if not is_user_logged_in() or session.get('role') != 'Organizer':
        return redirect('/login')

    f = _extract_event_form(request.form)
    error = validate_event_fields(
        f['title'], f['category'], f['event_date'],
        f['venue'], f['capacity'], f['registration_close_date']
    )
    if error:
        return redirect('/organizer_dashboard')

    try:
        cap = int(f['capacity'])
    except ValueError:
        return redirect('/organizer_dashboard')

    resubmit_event(
        event_id, session['user_id'], f['title'], f['category'],
        f['event_date'], f['event_time'], f['venue'], f['description'],
        cap, f['registration_close_date'] or None, f['contact_number']
    )

    evt = get_event_by_id(event_id)
    if evt:
        admin_user = get_user_by_email(ADMIN_EMAIL)
        if admin_user:
            create_notification(
                admin_user['id'],
                "Event Resubmitted for Approval",
                f"{session.get('user_name')} resubmitted '{evt['title']}' after fixing issues.",
                'info'
            )
        notify_admin_new_event(
            evt['title'], session.get('user_name', ''),
            evt['event_date'], evt['venue'], evt['event_code'] or ''
        )

    return redirect('/organizer_dashboard')


@app.route('/organizer/delete/<int:event_id>', methods=['POST'])
def delete_owned_event_route(event_id):
    if not is_user_logged_in() or session.get('role') != 'Organizer':
        return redirect('/login')
    delete_owned_event(event_id, session['user_id'])
    return redirect('/organizer_dashboard')


@app.route('/organizer/close_registration/<int:event_id>', methods=['POST'])
def close_registration_route(event_id):
    if not is_user_logged_in() or session.get('role') != 'Organizer':
        return redirect('/login')
    close_registration(event_id, session['user_id'])
    return redirect('/organizer_dashboard')


@app.route('/organizer/event_participants/<int:event_id>')
def organizer_event_participants(event_id):
    if not is_user_logged_in() or session.get('role') != 'Organizer':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403

    event = get_event_by_id(event_id)
    if not event or event['organizer_id'] != session['user_id']:
        return jsonify({'success': False, 'message': 'Event not found'}), 404

    participants = get_event_participants(event_id, session['user_id'])
    return jsonify({
        'success':     True,
        'event_title': event['title'],
        'event_code':  event['event_code'],
        'participants': [
            {
                'registration_id': p['registration_id'],
                'fullname':        p['fullname'],
                'email':           p['email'],
                'mobile':          p['mobile'],
                'registered_at':   str(p['registered_at']),
                'attended':        bool(p['attended']),
            }
            for p in participants
        ]
    }), 200


@app.route('/organizer/event_feedback/<int:event_id>')
def organizer_event_feedback(event_id):
    if not is_user_logged_in() or session.get('role') != 'Organizer':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403

    event = get_event_by_id(event_id)
    if not event or event['organizer_id'] != session['user_id']:
        return jsonify({'success': False, 'message': 'Event not found'}), 404

    rows = get_event_feedback(event_id, session['user_id'])
    avg_rating = round(sum(r['rating'] for r in rows) / len(rows), 1) if rows else None

    return jsonify({
        'success':     True,
        'event_title': event['title'],
        'avg_rating':  avg_rating,
        'count':       len(rows),
        'feedback': [
            {
                'fullname':   r['fullname'],
                'rating':     r['rating'],
                'comments':   r['comments'] or '',
                'created_at': str(r['created_at']),
            }
            for r in rows
        ]
    }), 200


@app.route('/organizer/mark_attendance/<int:registration_id>', methods=['POST'])
def mark_attendance_route(registration_id):
    if not is_user_logged_in() or session.get('role') != 'Organizer':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403

    data = request.get_json(silent=True) or {}
    attended_raw = data.get('attended', request.form.get('attended', 'true'))
    attended = attended_raw in (True, 'true', 'True', '1', 1)

    ok = mark_attendance(registration_id, session['user_id'], attended)
    if not ok:
        return jsonify({'success': False, 'message': 'Registration not found'}), 404
    return jsonify({'success': True, 'attended': attended}), 200


# ============= PARTICIPANT ACTIONS ============= #

@app.route('/participant/register_event/<int:event_id>', methods=['POST'])
def register_event_route(event_id):
    if not is_user_logged_in() or session.get('role') != 'Participant':
        return redirect('/login')

    result = register_for_event(event_id, session['user_id'])

    if result:
        evt  = get_event_by_id(event_id)
        user = get_user_by_id(session['user_id'])
        reg_id = result if isinstance(result, int) else 0

        if evt and user:
            # Email participant
            notify_participant_registered(
                user['email'], user['fullname'],
                evt['title'], evt['event_code'] or '',
                evt['event_date'], evt['venue'], reg_id
            )

            # In-app notification to participant
            create_notification(
                session['user_id'],
                "Registration Confirmed",
                f"You are registered for '{evt['title']}'. Your ID: REG-{reg_id:05d}",
                'success'
            )

            # In-app notification to organizer
            create_notification(
                evt['organizer_id'],
                "New Participant Registered",
                f"{user['fullname']} registered for your event '{evt['title']}'.",
                'info'
            )

    return redirect('/participant_dashboard')


@app.route('/participant/cancel_event/<int:event_id>', methods=['POST'])
def cancel_event_route(event_id):
    if not is_user_logged_in() or session.get('role') != 'Participant':
        return redirect('/login')

    evt  = get_event_by_id(event_id)
    user = get_user_by_id(session['user_id'])

    cancelled = cancel_registration(event_id, session['user_id'])

    if cancelled and evt and user:
        # Notify participant
        create_notification(
            session['user_id'],
            "Registration Cancelled",
            f"Your registration for '{evt['title']}' has been cancelled.",
            'warning'
        )
        # Notify organizer
        create_notification(
            evt['organizer_id'],
            "Participant Cancelled Registration",
            f"{user['fullname']} cancelled their registration for '{evt['title']}'.",
            'warning'
        )

    return redirect('/participant_dashboard')


# ============= ADMIN ACTIONS ============= #

@app.route('/admin/approve_event/<int:event_id>', methods=['POST'])
def approve_event_route(event_id):
    if not is_user_logged_in() or session.get('role') != 'Admin':
        return redirect('/login')

    event = get_event_by_id(event_id)
    approve_event(event_id)

    if event:
        organizer = _get_organizer_by_event(event)
        if organizer:
            notify_organizer_approved(
                organizer['email'], organizer['fullname'],
                event['title'], event['event_code'] or '',
                event['event_date']
            )
            create_notification(
                event['organizer_id'],
                "Event Approved",
                f"Your event '{event['title']}' has been approved and is now live!",
                'success'
            )

    return redirect('/admin_dashboard')


@app.route('/admin/reject_event/<int:event_id>', methods=['POST'])
def reject_event_route(event_id):
    if not is_user_logged_in() or session.get('role') != 'Admin':
        return redirect('/login')

    reason = request.form.get('reason', '').strip()
    if not reason:
        return redirect('/admin_dashboard')

    event = get_event_by_id(event_id)
    reject_event(event_id, reason)

    if event:
        organizer = _get_organizer_by_event(event)
        if organizer:
            notify_organizer_rejected(
                organizer['email'], organizer['fullname'],
                event['title'], reason
            )
            create_notification(
                event['organizer_id'],
                "Event Rejected",
                f"Your event '{event['title']}' was rejected. Reason: {reason}",
                'error'
            )

    return redirect('/admin_dashboard')


@app.route('/admin/bulk_approve', methods=['POST'])
def bulk_approve_route():
    if not is_user_logged_in() or session.get('role') != 'Admin':
        return redirect('/login')

    event_ids = [int(e) for e in request.form.getlist('event_ids')
                 if e.isdigit()]

    if event_ids:
        bulk_approve_events(event_ids)
        for event_id in event_ids:
            event = get_event_by_id(event_id)
            if event:
                organizer = _get_organizer_by_event(event)
                if organizer:
                    notify_organizer_approved(
                        organizer['email'], organizer['fullname'],
                        event['title'], event['event_code'] or '',
                        event['event_date']
                    )
                    create_notification(
                        event['organizer_id'],
                        "Event Approved",
                        f"Your event '{event['title']}' has been approved and is now live!",
                        'success'
                    )

    return redirect('/admin_dashboard')


@app.route('/admin/bulk_reject', methods=['POST'])
def bulk_reject_route():
    if not is_user_logged_in() or session.get('role') != 'Admin':
        return redirect('/login')

    reason    = request.form.get('bulk_reason', '').strip()
    event_ids = [int(e) for e in request.form.getlist('event_ids')
                 if e.isdigit()]

    if not reason:
        return redirect('/admin_dashboard')

    if event_ids:
        bulk_reject_events(event_ids, reason)
        for event_id in event_ids:
            event = get_event_by_id(event_id)
            if event:
                organizer = _get_organizer_by_event(event)
                if organizer:
                    notify_organizer_rejected(
                        organizer['email'], organizer['fullname'],
                        event['title'], reason
                    )
                    create_notification(
                        event['organizer_id'],
                        "Event Rejected",
                        f"Your event '{event['title']}' was rejected. Reason: {reason}",
                        'error'
                    )

    return redirect('/admin_dashboard')


@app.route('/admin/delete_user/<int:user_id>', methods=['POST'])
def admin_delete_user_route(user_id):
    if not is_user_logged_in() or session.get('role') != 'Admin':
        return redirect('/login')
    delete_user(user_id)
    return redirect('/admin_dashboard')


@app.route('/admin/delete_event/<int:event_id>', methods=['POST'])
def admin_delete_event_route(event_id):
    if not is_user_logged_in() or session.get('role') != 'Admin':
        return redirect('/login')
    admin_delete_event(event_id)
    return redirect('/admin_dashboard')


@app.route('/admin/trust_organizer/<int:organizer_id>', methods=['POST'])
def trust_organizer_route(organizer_id):
    """Toggle trusted status for an organizer — auto-approves future events."""
    if not is_user_logged_in() or session.get('role') != 'Admin':
        return redirect('/login')
    trusted = request.form.get('trusted', 'true').lower() == 'true'
    set_organizer_trusted(organizer_id, trusted)
    if trusted:
        create_notification(
            organizer_id,
            "Trusted Organizer Status Granted",
            "Your future events will be auto-approved without manual review.",
            'success'
        )
    return redirect('/admin_dashboard')


@app.route('/admin/contact/<int:message_id>/read', methods=['POST'])
def mark_contact_read_route(message_id):
    if not is_user_logged_in() or session.get('role') != 'Admin':
        return redirect('/login')
    mark_contact_read(message_id)
    return redirect('/admin_dashboard')


@app.route('/admin/contact/<int:message_id>/reply', methods=['POST'])
def reply_contact_route(message_id):
    if not is_user_logged_in() or session.get('role') != 'Admin':
        return redirect('/login')

    reply_text = (request.form.get('reply') or '').strip()
    if not reply_text:
        return redirect('/admin_dashboard')

    msg = get_contact_message_by_id(message_id)
    if not msg:
        return redirect('/admin_dashboard')

    reply_to_contact_message(message_id, reply_text)
    notify_contact_reply(msg['email'], msg['full_name'], msg['subject'],
                         msg['message'], reply_text)

    return redirect('/admin_dashboard')


# ============= DOWNLOADABLE REPORTS ============= #

@app.route('/participant/receipt/<int:registration_id>')
def download_receipt(registration_id):
    if not is_user_logged_in() or session.get('role') != 'Participant':
        return redirect('/login')

    reg = get_registration_by_id(registration_id, session['user_id'])
    if not reg:
        return redirect('/participant_dashboard')

    buffer = BytesIO()
    doc    = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    elements = [
        Paragraph("EventHub — Registration Receipt", styles['Title']),
        Spacer(1, 16),
    ]
    rows = [
        ['Registration ID', f"REG-{reg['registration_id']:05d}"],
        ['Event Code',      reg['event_code'] or '-'],
        ['Event Title',     reg['event_title']],
        ['Category',        reg['category']],
        ['Date',            reg['event_date']],
        ['Time',            reg['event_time'] or '-'],
        ['Venue',           reg['venue']],
        ['Organizer',       reg['organizer_name']],
        ['Participant',     reg['participant_name']],
        ['Email',           reg['participant_email']],
        ['Registered On',   str(reg['registered_at'])],
    ]
    tbl = Table(rows, colWidths=[160, 320])
    tbl.setStyle(TableStyle([
        ('FONTNAME',      (0, 0), (0, -1), 'Helvetica-Bold'),
        ('GRID',          (0, 0), (-1, -1), 0.5, colors.grey),
        ('TOPPADDING',    (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING',   (0, 0), (-1, -1), 6),
        ('RIGHTPADDING',  (0, 0), (-1, -1), 6),
        ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(tbl)
    doc.build(elements)
    buffer.seek(0)

    filename = f"receipt_{reg['event_code'] or registration_id}.pdf"
    return send_file(buffer, as_attachment=True,
                     download_name=filename, mimetype='application/pdf')


@app.route('/participant/certificate/<int:registration_id>')
def download_certificate(registration_id):
    if not is_user_logged_in() or session.get('role') != 'Participant':
        return redirect('/login')

    reg = get_registration_by_id(registration_id, session['user_id'])
    if not reg:
        return redirect('/participant_dashboard')

    # Certificates require: event completed, attendance marked by the
    # organizer, and feedback submitted by the participant.
    if reg['event_status'] != 'Completed' or not reg['attended'] or not reg['feedback_submitted']:
        return redirect('/participant_dashboard')

    buffer = build_certificate_pdf(reg)
    filename = f"certificate_{reg['event_code'] or registration_id}.pdf"
    return send_file(buffer, as_attachment=True,
                     download_name=filename, mimetype='application/pdf')


@app.route('/participant/feedback/<int:registration_id>', methods=['POST'])
def submit_feedback_route(registration_id):
    if not is_user_logged_in() or session.get('role') != 'Participant':
        return redirect('/login')

    try:
        rating = int(request.form.get('rating', 0))
    except ValueError:
        rating = 0
    comments = (request.form.get('comments') or '').strip()

    if rating < 1 or rating > 5:
        return redirect('/participant_dashboard')

    reg = get_registration_by_id(registration_id, session['user_id'])
    if not reg:
        return redirect('/participant_dashboard')

    submit_feedback(registration_id, reg['event_id'], session['user_id'], rating, comments)
    return redirect('/participant_dashboard')


@app.route('/participant/qr/<int:registration_id>')
def participant_checkin_qr(registration_id):
    """
    Returns a QR code PNG (not a download) that encodes a signed check-in
    link for this registration. Participants show this at the event
    entrance; the organizer scans it with their phone's normal camera app,
    which opens the check-in link in a browser and marks them present.
    """
    if not is_user_logged_in() or session.get('role') != 'Participant':
        return redirect('/login')

    reg = get_registration_by_id(registration_id, session['user_id'])
    if not reg:
        return redirect('/participant_dashboard')

    token = build_checkin_token(registration_id)
    checkin_url = request.host_url.rstrip('/') + f'/organizer/checkin/{registration_id}/{token}'
    buffer = build_qr_png(checkin_url)
    return send_file(buffer, mimetype='image/png')


@app.route('/organizer/checkin/<int:registration_id>/<token>')
def organizer_checkin(registration_id, token):
    """
    Reached by scanning a participant's QR code. Verifies the token, then
    marks that registration as attended for the organizer's own event.
    """
    if not is_user_logged_in() or session.get('role') != 'Organizer':
        return redirect('/login')

    expected_token = build_checkin_token(registration_id)
    if not hmac.compare_digest(token, expected_token):
        return "<h2 style='font-family:sans-serif;color:#ef6361;'>Invalid or tampered check-in code.</h2>", 400

    ok = mark_attendance(registration_id, session['user_id'], True)
    if not ok:
        return ("<h2 style='font-family:sans-serif;color:#ef6361;'>"
                "This registration doesn't belong to one of your events, "
                "or doesn't exist.</h2>"), 404

    return redirect('/organizer_dashboard?checked_in=' + str(registration_id))


@app.route('/organizer/report/event/<int:event_id>')
def organizer_event_report(event_id):
    if not is_user_logged_in() or session.get('role') != 'Organizer':
        return redirect('/login')

    event = get_event_by_id(event_id)
    if not event or event['organizer_id'] != session['user_id']:
        return redirect('/organizer_dashboard')

    fmt          = request.args.get('format', 'pdf')
    participants = get_event_participants(event_id, session['user_id'])
    event_code   = event['event_code'] or f"EVT-{event_id:04d}"
    headers = ['Reg. ID', 'Name', 'Email', 'Mobile', 'Registered On']
    rows = [
        [f"REG-{p['registration_id']:05d}", p['fullname'],
         p['email'], p['mobile'], str(p['registered_at'])]
        for p in participants
    ]

    if fmt == 'excel':
        buf = build_excel_table(headers, rows)
        return send_file(buf, as_attachment=True,
                         download_name=f"{event_code}_participants.xlsx",
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    buf = build_pdf_table(
        f"{event['title']} ({event_code}) — Registered Participants",
        headers, rows
    )
    return send_file(buf, as_attachment=True,
                     download_name=f"{event_code}_participants.pdf",
                     mimetype='application/pdf')


@app.route('/organizer/report/overall')
def organizer_overall_report():
    if not is_user_logged_in() or session.get('role') != 'Organizer':
        return redirect('/login')

    fmt           = request.args.get('format', 'pdf')
    registrations = get_all_registrations_by_organizer(session['user_id'])
    headers = ['Reg. ID', 'Event Code', 'Event', 'Category',
               'Date', 'Participant', 'Email', 'Mobile', 'Registered On']
    rows = [
        [f"REG-{r['registration_id']:05d}", r['event_code'], r['event_title'],
         r['category'], r['event_date'], r['fullname'],
         r['email'], r['mobile'], str(r['registered_at'])]
        for r in registrations
    ]

    if fmt == 'excel':
        buf = build_excel_table(headers, rows)
        return send_file(buf, as_attachment=True,
                         download_name="overall_registrations.xlsx",
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    buf = build_pdf_table("Overall Registered Participants Report", headers, rows)
    return send_file(buf, as_attachment=True,
                     download_name="overall_registrations.pdf",
                     mimetype='application/pdf')


@app.route('/admin/report/all')
def admin_all_report():
    if not is_user_logged_in() or session.get('role') != 'Admin':
        return redirect('/login')

    fmt           = request.args.get('format', 'pdf')
    registrations = get_all_registrations_admin()
    headers = ['Reg. ID', 'Event Code', 'Event', 'Category', 'Date',
               'Organizer', 'Participant', 'Email', 'Mobile', 'Registered On']
    rows = [
        [f"REG-{r['registration_id']:05d}", r['event_code'], r['event_title'],
         r['category'], r['event_date'], r['organizer_name'],
         r['fullname'], r['email'], r['mobile'], str(r['registered_at'])]
        for r in registrations
    ]

    if fmt == 'excel':
        buf = build_excel_table(headers, rows)
        return send_file(buf, as_attachment=True,
                         download_name="system_registrations.xlsx",
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    buf = build_pdf_table("System-wide Registered Participants Report", headers, rows)
    return send_file(buf, as_attachment=True,
                     download_name="system_registrations.pdf",
                     mimetype='application/pdf')


# ============= LOGOUT ============= #

@app.route('/logout')
def logout():
    session.clear()
    return jsonify({'success': True,
                    'message': 'Logged out successfully!',
                    'redirect': '/'}), 200


# ============= ERROR HANDLERS ============= #

@app.errorhandler(404)
def page_not_found(e):
    return render_template('404.html'), 404


@app.errorhandler(500)
def internal_error(e):
    return render_template('500.html'), 500


@app.errorhandler(429)
def rate_limit_exceeded(e):
    # login/register are called via fetch() expecting JSON — a plain-text
    # 429 would break their .json() parsing on the frontend.
    return jsonify({'success': False,
                    'message': 'Too many attempts. Please wait a moment and try again.'}), 429


@app.errorhandler(400)
def bad_request(e):
    # Covers CSRF validation failures among other 400s. Forms posted from
    # this app always include the token, so this mainly protects against
    # forged/expired requests from elsewhere.
    if request.path.startswith(('/login', '/register', '/contact')) or request.is_json:
        return jsonify({'success': False,
                        'message': 'Your session expired or the request was invalid. Please refresh the page and try again.'}), 400
    return render_template('500.html'), 400


# ============= RUN ============= #

if __name__ == '__main__':
    init_db()
    print("🚀 EVENTHUB SERVER STARTING...")
    app.run(debug=False, host='0.0.0.0',
            port=int(os.environ.get('PORT', 5000)))